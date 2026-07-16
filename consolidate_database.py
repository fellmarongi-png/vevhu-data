import csv
import glob
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

repo_dir = '/home/ec2-user/vevhu-data'
batch_files = sorted(glob.glob(os.path.join(repo_dir, '**/batch_*.csv'), recursive=True))

all_records = []

print(f"Found batch files for consolidation: {[os.path.relpath(f, repo_dir) for f in batch_files]}")

# We will also keep track of records by batch to write to individual sheets
records_by_batch = {}

for b_file in batch_files:
    batch_name = os.path.basename(b_file)
    batch_key = batch_name.replace('.csv', '')
    records_by_batch[batch_key] = []
    
    with open(b_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader, start=1):
            get_val = lambda k: (row.get(k) or '').strip()
            
            file_name = get_val('File Name')
            stand_no = get_val('Stand No')
            name = get_val('Name & Surname')
            id_no = get_val('ID No')
            cell_no = get_val('Cell No')
            comp_yes = get_val('Compliance Yes')
            comp_no = get_val('Compliance No')
            comments = get_val('Comments')
            
            # Determine Record Status
            is_blank = not (stand_no or name or id_no or cell_no or comments)
            is_crossed_out = '[CROSSED OUT]' in comments or 'crossed out' in comments.lower() or '[crossed out]' in comments.lower()
            
            if is_blank:
                status = 'Blank Row'
            elif is_crossed_out:
                status = 'Crossed Out'
            else:
                status = 'Active'
                
            # Cleaned Stand No
            cleaned_stand = re.sub(r'\s+', ' ', stand_no).strip()
            
            # Cleaned Name
            cleaned_name = re.sub(r'\s+', ' ', name).strip()
            if cleaned_name.endswith('.'):
                cleaned_name = cleaned_name[:-1].strip()
                
            # Cleaned ID No
            cleaned_id = re.sub(r'\s+', ' ', id_no).strip()
            
            # Cleaned Cell No
            cleaned_cell = re.sub(r'\s+', ' ', cell_no).strip()
            
            rec = {
                'File Name': file_name,
                'Stand No': cleaned_stand,
                'Name & Surname': cleaned_name,
                'ID No': cleaned_id,
                'Cell No': cleaned_cell,
                'Compliance Yes': comp_yes,
                'Compliance No': comp_no,
                'Comments': comments,
                'Batch Source': batch_name,
                'Row Index': str(idx),
                'Record Status': status
            }
            all_records.append(rec)
            records_by_batch[batch_key].append({
                'File Name': file_name,
                'Stand No': cleaned_stand,
                'Name & Surname': cleaned_name,
                'ID No': cleaned_id,
                'Cell No': cleaned_cell,
                'Compliance Yes': comp_yes,
                'Compliance No': comp_no,
                'Comments': comments
            })

print(f'Total records read: {len(all_records)}')

# Write master_vevhu_database.csv
master_csv_path = os.path.join(repo_dir, 'master_vevhu_database.csv')
if all_records:
    fieldnames = list(all_records[0].keys())
    with open(master_csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_records)
    print(f'Saved master database to {master_csv_path}')
else:
    print('Error: No records found to write!')
    exit(1)

# Generate formatted Excel Workbook
wb = Workbook()

# Styling tokens
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Light grayish blue
border_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
align_left = Alignment(horizontal="left", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

def format_sheet(ws, headers, rows):
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Freeze pane
    ws.freeze_panes = "A2"
    
    # Write header
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center if col_idx != len(headers) else align_left
        cell.border = thin_border
    
    # Write rows
    for r_idx, row in enumerate(rows, start=2):
        ws.append(row)
        is_even = (r_idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = align_left
            if is_even:
                cell.fill = zebra_fill
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# 1. Write the Consolidated Master Sheet
ws_master = wb.active
ws_master.title = "Master Database"
master_headers = list(all_records[0].keys())
master_rows = [list(r.values()) for r in all_records]
format_sheet(ws_master, master_headers, master_rows)

# 2. Write Individual Batch Sheets
# Sort batches so sheets are added in order
for b_key in sorted(records_by_batch.keys()):
    # Sheet name max length in Excel is 31 characters
    sheet_name = b_key.replace('batch_', '').replace('_', ' ')[:31]
    ws = wb.create_sheet(title=sheet_name)
    
    batch_headers = ['File Name', 'Stand No', 'Name & Surname', 'ID No', 'Cell No', 'Compliance Yes', 'Compliance No', 'Comments']
    batch_rows = [list(r.values()) for r in records_by_batch[b_key]]
    format_sheet(ws, batch_headers, batch_rows)

# Save the workbook
master_xlsx_path = os.path.join(repo_dir, 'master_vevhu_database.xlsx')
wb.save(master_xlsx_path)
print(f"Saved styled Excel workbook to {master_xlsx_path}")
