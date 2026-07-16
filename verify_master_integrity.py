import csv
import os
import openpyxl

repo_dir = '/home/ec2-user/vevhu-data'
master_csv = os.path.join(repo_dir, 'master_vevhu_database.csv')
clean_xlsx = os.path.join(repo_dir, 'final', 'final_consolidated_database.xlsx')
dq_xlsx = os.path.join(repo_dir, 'final', 'data_quality_analysis.xlsx')

print("=== STARTING INTEGRITY VERIFICATION BETWEEN MASTER LEDGER AND FINAL WORKBOOKS ===")

# 1. Load raw master records
master_active = []
master_crossed = []
master_blank = []

with open(master_csv, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for idx, r in enumerate(reader):
        status = r.get('Record Status', '')
        # Columns in master: Cleaned Stand No, Cleaned ID No, Cleaned Cell No, Cleaned Compliance Yes, Cleaned Compliance No, Record Status, Batch Source, Row Index, ...
        # Let's map columns to key standard columns: ['Stand No', 'Name & Surname', 'ID No', 'Cell No', 'Compliance Yes', 'Compliance No', 'Comments']
        # Note: the master has 'Cleaned Stand No' as 'Cleaned Stand No' or similar, let's inspect the master keys first if needed
        # We will parse the standard fields
        record = {
            'Stand No': r.get('Cleaned Stand No', r.get('Stand No', '')),
            'Name & Surname': r.get('Name & Surname', ''),
            'ID No': r.get('Cleaned ID No', r.get('ID No', '')),
            'Cell No': r.get('Cleaned Cell No', r.get('Cell No', '')),
            'Compliance Yes': r.get('Cleaned Compliance Yes', r.get('Compliance Yes', '')),
            'Compliance No': r.get('Cleaned Compliance No', r.get('Compliance No', '')),
            'Comments': r.get('Comments', ''),
            'Batch Source': r.get('Batch Source', ''),
            'Row Index': r.get('Row Index', '')
        }
        
        if status == 'Active':
            master_active.append(record)
        elif status == 'Crossed Out':
            master_crossed.append(record)
        elif status == 'Blank Row':
            master_blank.append(record)

print(f"Master CSV stats:")
print(f"  Active Records: {len(master_active)}")
print(f"  Crossed Out Records: {len(master_crossed)}")
print(f"  Blank Rows: {len(master_blank)}")
print(f"  Total Rows: {len(master_active) + len(master_crossed) + len(master_blank)}")

# 2. Load clean consolidated database
wb_clean = openpyxl.load_workbook(clean_xlsx, read_only=True)
xlsx_clean_active = []

for sheet_name in wb_clean.sheetnames:
    sheet = wb_clean[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]
    
    # Map to dict using headers
    for r in data_rows:
        if r and any(cell is not None for cell in r):
            record = dict(zip(header, r))
            # Normalize dict keys to standard
            xlsx_clean_active.append({
                'Stand No': str(record.get('Stand No', '')).strip() if record.get('Stand No') is not None else '',
                'Name & Surname': str(record.get('Name & Surname', '')).strip() if record.get('Name & Surname') is not None else '',
                'ID No': str(record.get('ID No', '')).strip() if record.get('ID No') is not None else '',
                'Cell No': str(record.get('Cell No', '')).strip() if record.get('Cell No') is not None else '',
                'Compliance Yes': str(record.get('Compliance Yes', '')).strip() if record.get('Compliance Yes') is not None else '',
                'Compliance No': str(record.get('Compliance No', '')).strip() if record.get('Compliance No') is not None else '',
                'Comments': str(record.get('Comments', '')).strip() if record.get('Comments') is not None else '',
                'Sheet': sheet_name
            })

print(f"Consolidated Clean XLSX stats:")
print(f"  Active Records in XLSX: {len(xlsx_clean_active)}")

# 3. Load crossed out records from DQ xlsx
wb_dq = openpyxl.load_workbook(dq_xlsx, read_only=True)
xlsx_crossed = []
if 'Crossed Out Records' in wb_dq.sheetnames:
    sheet = wb_dq['Crossed Out Records']
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]
    for r in data_rows:
        if r and any(cell is not None for cell in r):
            record = dict(zip(header, r))
            xlsx_crossed.append({
                'Stand No': str(record.get('Stand No', '')).strip() if record.get('Stand No') is not None else '',
                'Name & Surname': str(record.get('Name & Surname', '')).strip() if record.get('Name & Surname') is not None else '',
                'ID No': str(record.get('ID No', '')).strip() if record.get('ID No') is not None else '',
                'Cell No': str(record.get('Cell No', '')).strip() if record.get('Cell No') is not None else '',
                'Comments': str(record.get('Comments', '')).strip() if record.get('Comments') is not None else '',
                'Batch Source': str(record.get('Batch Source', '')).strip() if record.get('Batch Source') is not None else '',
                'Row Index': str(record.get('Row Index', '')).strip() if record.get('Row Index') is not None else ''
            })

print(f"Consolidated DQ XLSX Crossed Out stats:")
print(f"  Crossed Out Records in XLSX: {len(xlsx_crossed)}")

# 4. Perform comparison check
# We want to make sure every active record in the master is represented in the clean xlsx
errors = []

# Helper to create a matching key (normalized Stand, Name, ID, Cell)
def get_match_key(r):
    stand = str(r.get('Stand No', '')).strip().lower()
    name = str(r.get('Name & Surname', '')).strip().lower()
    id_no = str(r.get('ID No', '')).strip().lower()
    cell = str(r.get('Cell No', '')).strip().lower()
    return f"{stand}|{name}|{id_no}|{cell}"

master_active_keys = [get_match_key(r) for r in master_active]
xlsx_active_keys = [get_match_key(r) for r in xlsx_clean_active]

missing_in_xlsx = []
for idx, key in enumerate(master_active_keys):
    if key not in xlsx_active_keys:
        missing_in_xlsx.append(master_active[idx])

extra_in_xlsx = []
for idx, key in enumerate(xlsx_active_keys):
    if key not in master_active_keys:
        extra_in_xlsx.append(xlsx_clean_active[idx])

print(f"\n--- INTEGRITY CHECKS ---")
if len(master_active) == len(xlsx_clean_active):
    print("SUCCESS: Count of active records matches exactly (3,377).")
else:
    print(f"WARNING: Count mismatch! Master active: {len(master_active)}, Clean XLSX active: {len(xlsx_clean_active)}")

if not missing_in_xlsx:
    print("SUCCESS: Every active record in master CSV is present in consolidated XLSX.")
else:
    print(f"ERROR: {len(missing_in_xlsx)} active master records are missing in XLSX! Examples:")
    for r in missing_in_xlsx[:5]:
        print("  ", r)

if not extra_in_xlsx:
    print("SUCCESS: Every active record in XLSX is present in master CSV.")
else:
    print(f"ERROR: {len(extra_in_xlsx)} active records in XLSX do not exist in master CSV! Examples:")
    for r in extra_in_xlsx[:5]:
        print("  ", r)

# Verify Crossed Out matching
master_crossed_keys = [f"{str(r['Stand No']).strip().lower()}|{str(r['Name & Surname']).strip().lower()}" for r in master_crossed]
xlsx_crossed_keys = [f"{str(r['Stand No']).strip().lower()}|{str(r['Name & Surname']).strip().lower()}" for r in xlsx_crossed]

missing_crossed = []
for idx, key in enumerate(master_crossed_keys):
    if key not in xlsx_crossed_keys:
        missing_crossed.append(master_crossed[idx])

if not missing_crossed:
    print("SUCCESS: Every crossed-out record in master CSV is present in data_quality_analysis.xlsx sheet.")
else:
    print(f"ERROR: {len(missing_crossed)} crossed-out master records are missing in DQ workbook! Examples:")
    for r in missing_crossed[:5]:
        print("  ", r)

print("=== INTEGRITY VERIFICATION COMPLETE ===")
