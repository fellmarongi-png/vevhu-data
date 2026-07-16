import csv
import os
import re
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

repo_dir = '/home/ec2-user/vevhu-data'
master_csv = os.path.join(repo_dir, 'master_vevhu_database.csv')
final_dir = os.path.join(repo_dir, 'final')
os.makedirs(final_dir, exist_ok=True)

# Styling tokens for professional sheets
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Light grayish blue
border_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
align_left = Alignment(horizontal="left", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

def format_sheet(ws, headers, rows):
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"
    
    # Write headers
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center if col_idx != len(headers) else align_left
        cell.border = thin_border
        
    # Write rows
    for r_idx, row in enumerate(rows, start=2):
        ws.append(row)
        is_even = (r_idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = align_left
            if is_even:
                cell.fill = zebra_fill
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Helper function to validate phone format
def is_valid_phone(cell):
    if not cell:
        return True
    # Strip spaces, dashes, parentheses
    clean = cell.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')
    
    # Standard local 10-digit e.g., 0778812565
    if len(clean) == 10 and clean.startswith('0'):
        return clean.isdigit()
    # Standard international e.g., 263778812565
    if len(clean) == 12 and clean.startswith('263'):
        return clean.isdigit()
    # Standard local 9-digit e.g., 778812565
    if len(clean) == 9 and (clean.startswith('7') or clean.startswith('8') or clean.startswith('1')):
        return clean.isdigit()
    return False

def validate_phone_field(cell_field):
    if not cell_field:
        return True
    parts = re.split(r'/|,|&', cell_field)
    for p in parts:
        p = p.strip()
        if p and not is_valid_phone(p):
            return False
    return True

# Helper function to check Zimbabwean Passport format (2 letters followed by 6 digits)
def is_passport_format(id_val):
    if not id_val:
        return False
    clean = id_val.replace(' ', '').replace('-', '').upper()
    return len(clean) == 8 and clean[:2].isalpha() and clean[2:].isdigit()

# Helper function to check standard Zimbabwean ID format
def is_valid_national_id(id_val):
    if not id_val:
        return True
    clean = id_val.replace(' ', '').replace('-', '').upper()
    # Standard National ID: 11-111111X11 -> clean has 10 or 11 characters (digits + letter + digits)
    return bool(re.match(r'^\d{8,10}[A-Z]\d{2}$', clean))

# Helper function to identify missing/placeholder names
def is_missing_name(name):
    if not name:
        return True
    n = str(name).strip().lower()
    n = re.sub(r'\s+', ' ', n)
    if n in ['', 'no details', 'nodetails', 'nodeals', 'no details.', 'nil', 'none', 'no information', 'no name', 'no occupant', 'unoccupied', 'no cabin']:
        return True
    # check for patterns like "no details 1234"
    if re.match(r'^(no details|nodetails|no occupant|unoccupied)\s*\d*$', n):
        return True
    return False

# Read master records
active_records = []
all_records = []

with open(master_csv, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        all_records.append(row)
        if row['Record Status'] == 'Active':
            active_records.append(row)

# ----------------------------------------------------
# A. BUILD CLEAN CONSOLIDATED WORKBOOK (final_consolidated_database.xlsx)
# ----------------------------------------------------
wb_clean = Workbook()
# Remove default sheet
default_sheet = wb_clean.active
wb_clean.remove(default_sheet)

clean_headers = ['Stand No', 'Name & Surname', 'ID No', 'Cell No', 'Compliance Yes', 'Compliance No', 'Comments']

# Group active records by category
categorized_data = {
    '200m2': [],
    '300m2': [],
    'Madzimai': [],
    'Kambarami': [],
    'Non-Standard': []
}

for row in active_records:
    b_src = row['Batch Source'].lower()
    comments = row['Comments'].lower()
    
    # Extract clean comments: if it was a tablet record, map only Notes to Comments, removing money and lot columns
    comments_clean = row['Comments']
    if 'tablet' in b_src:
        # Check if we can extract Notes from the formatted comments string
        match = re.search(r'Notes:\s*([^,]+)', row['Comments'])
        if match:
            comments_clean = match.group(1).strip()
        else:
            # If no Notes label, check if comment starts with Size/Lot and remove them
            # if only size and lot exist, leave comments blank
            if row['Comments'].startswith('Size:') or row['Comments'].startswith('Lot:'):
                notes_match = re.search(r'Notes:\s*(.*)$', row['Comments'])
                if notes_match:
                    comments_clean = notes_match.group(1).strip()
                else:
                    comments_clean = ''
                    
    # Format the data row
    clean_row = [
        row['Stand No'],
        row['Name & Surname'],
        row['ID No'],
        row['Cell No'],
        row['Compliance Yes'],
        row['Compliance No'],
        comments_clean
    ]
    
    # Categorization logic
    if 'madzimai' in b_src:
        categorized_data['Madzimai'].append(clean_row)
    elif 'kambarami' in b_src:
        categorized_data['Kambarami'].append(clean_row)
    elif 'non_standard' in b_src:
        categorized_data['Non-Standard'].append(clean_row)
    elif '300m2' in b_src or 'tablet_300' in b_src or 'tat_300' in b_src or ('magama' in b_src and '300' in comments):
        categorized_data['300m2'].append(clean_row)
    else:
        categorized_data['200m2'].append(clean_row)

# Create sheets in order
for category in ['200m2', '300m2', 'Madzimai', 'Kambarami', 'Non-Standard']:
    # Replace suffix or name for final sheet display
    sheet_name = category.replace('m2', 'm²')
    ws = wb_clean.create_sheet(title=sheet_name)
    format_sheet(ws, clean_headers, categorized_data[category])
    print(f"Clean database sheet: {sheet_name} ({len(categorized_data[category])} active records)")

clean_xlsx_path = os.path.join(final_dir, 'final_consolidated_database.xlsx')
wb_clean.save(clean_xlsx_path)
print(f"Saved clean database workbook to {clean_xlsx_path}")

# ----------------------------------------------------
# B. BUILD DATA QUALITY WORKBOOK (data_quality_analysis.xlsx)
# ----------------------------------------------------
wb_audit = Workbook()
default_sheet_audit = wb_audit.active
wb_audit.remove(default_sheet_audit)

audit_headers = ['Stand No', 'Name & Surname', 'ID No', 'Cell No', 'Comments', 'Batch Source', 'Row Index']

# 1. Dual Registrations
# Stands of the same size group containing multiple owners
stand_owners = defaultdict(list)
for row in active_records:
    stand = row['Stand No'].strip()
    name = row['Name & Surname'].strip()
    b_src = row['Batch Source'].lower()
    comments = row['Comments'].lower()
    
    if stand and stand not in ['Missing', 'Kambarami', 'Madzimai', '0', '']:
        # Determine stand size group
        if 'madzimai' in b_src:
            size_group = 'Madzimai'
        elif 'kambarami' in b_src:
            size_group = 'Kambarami'
        elif 'non_standard' in b_src:
            size_group = 'Non-Standard'
        elif '300m2' in b_src or 'tablet_300' in b_src or 'tat_300' in b_src or ('magama' in b_src and '300' in comments):
            size_group = '300m2'
        else:
            size_group = '200m2'
            
        key = (stand, size_group)
        stand_owners[key].append(row)

dual_rows = []
for (stand, size), rows in sorted(stand_owners.items(), key=lambda x: (x[0][1], x[0][0])):
    unique_names = set(r['Name & Surname'].strip().lower() for r in rows if r['Name & Surname'])
    if len(unique_names) > 1:
        for r in rows:
            dual_rows.append([
                f"{r['Stand No']} ({size})",
                r['Name & Surname'],
                r['ID No'],
                r['Cell No'],
                r['Comments'],
                r['Batch Source'],
                r['Row Index']
            ])

# 2. Passport IDs Used
passport_rows = []
for r in active_records:
    if is_passport_format(r['ID No']):
        passport_rows.append([
            r['Stand No'], r['Name & Surname'], r['ID No'], r['Cell No'],
            r['Comments'], r['Batch Source'], r['Row Index']
        ])

# 3. Invalid Phone Numbers
invalid_phone_rows = []
for r in active_records:
    if r['Cell No'] and not validate_phone_field(r['Cell No']):
        invalid_phone_rows.append([
            r['Stand No'], r['Name & Surname'], r['ID No'], r['Cell No'],
            r['Comments'], r['Batch Source'], r['Row Index']
        ])

# 4. No Names
no_name_rows = []
for r in active_records:
    name_val = r['Name & Surname']
    if is_missing_name(name_val) and (r['Stand No'] or r['Cell No']):
        display_name = name_val if name_val else '[No Name Captured]'
        no_name_rows.append([
            r['Stand No'], display_name, r['ID No'], r['Cell No'],
            r['Comments'], r['Batch Source'], r['Row Index']
        ])

# 5. No Stand Numbers
no_stand_rows = []
for r in active_records:
    if not r['Stand No'] or r['Stand No'].lower() in ['missing', '0', '']:
        no_stand_rows.append([
            'Missing', r['Name & Surname'], r['ID No'], r['Cell No'],
            r['Comments'], r['Batch Source'], r['Row Index']
        ])

# 6. Non-Standard National IDs
non_std_id_rows = []
for r in active_records:
    id_val = r['ID No']
    if id_val and not is_passport_format(id_val) and not is_valid_national_id(id_val):
        non_std_id_rows.append([
            r['Stand No'], r['Name & Surname'], r['ID No'], r['Cell No'],
            r['Comments'], r['Batch Source'], r['Row Index']
        ])

# 7. Crossed Out Records (Fully cancelled entries from original sheets)
crossed_out_rows = []
for r in all_records:
    if r['Record Status'] == 'Crossed Out':
        crossed_out_rows.append([
            r['Stand No'], r['Name & Surname'], r['ID No'], r['Cell No'],
            r['Comments'], r['Batch Source'], r['Row Index']
        ])

# 8. Property Structures from Comments
structure_rows = defaultdict(list)
for r in active_records:
    com_lower = r['Comments'].lower()
    row_data = [
        r['Stand No'], r['Name & Surname'], r['ID No'], r['Cell No'],
        r['Comments'], r['Batch Source'], r['Row Index']
    ]
    if 'cabin' in com_lower:
        structure_rows['Cabins'].append(row_data)
    elif 'cottage' in com_lower:
        structure_rows['Cottages'].append(row_data)
    elif 'unfinished' in com_lower or 'foundation' in com_lower:
        structure_rows['Unfinished Structures'].append(row_data)
    elif 'tuckshop' in com_lower:
        structure_rows['Tuckshops'].append(row_data)
    elif 'double' in com_lower or 'displaced' in com_lower:
        structure_rows['Double Allocations'].append(row_data)

# Create Audit worksheets
sheets_to_create = [
    ('Dual Registrations', dual_rows),
    ('Passport IDs Used', passport_rows),
    ('Invalid Phone Numbers', invalid_phone_rows),
    ('No Names', no_name_rows),
    ('No Stand Numbers', no_stand_rows),
    ('Non-Std National IDs', non_std_id_rows),
    ('Crossed Out Records', crossed_out_rows)
]

for title, data in sheets_to_create:
    ws = wb_audit.create_sheet(title=title)
    format_sheet(ws, audit_headers, data)
    print(f"Audit sheet: {title} ({len(data)} records)")

# Create structure sheets
for title, data in sorted(structure_rows.items()):
    ws = wb_audit.create_sheet(title=title)
    format_sheet(ws, audit_headers, data)
    print(f"Structure sheet: {title} ({len(data)} records)")

audit_xlsx_path = os.path.join(final_dir, 'data_quality_analysis.xlsx')
wb_audit.save(audit_xlsx_path)
print(f"Saved audit reports workbook to {audit_xlsx_path}")

print("All reports compiled successfully!")
